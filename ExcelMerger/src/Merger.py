import openpyxl,os

class BreedType:
    HF = 0
    Jy = 0
    ND = 0
    total = 0

class HMBType:
    name = ''
    unbred = BreedType()
    bred = BreedType()
    pregnant = BreedType()
    milkYield = BreedType()

class FullRecord:
    id = ''
    age = ''
    breed = ''
    status = ''
    breedingStatus = ''
    sinceLastAI = ''
    daysInMilk = ''
    toConception = ''
    avgMilk = ''
    routeName = ''
    hmbName = ''

print('Loading the configuration')
configBook = openpyxl.load_workbook(os.path.join('..','Config','Config.xlsx'))
configSheet = configBook.get_active_sheet()
mergeColumns = configSheet['A'][1:];
firstColumn = mergeColumns[0]

print('Loading the Breed configuration')
breedConfigBook = openpyxl.load_workbook(os.path.join('..','Config','BreedConfig.xlsx'))
breedConfigSheet = breedConfigBook.get_active_sheet()
breedRows = breedConfigSheet[2:breedConfigSheet.max_row]
breedMap = {}
for breedRow in breedRows:
    breedMap[breedRow[0].value] = breedRow[1].value

outputBook = openpyxl.Workbook()
outputSheet = outputBook.get_active_sheet()
mergeColList = []
for mergeCol in mergeColumns:
    mergeColList.append(mergeCol.value)
outputSheet.append(mergeColList)

def validateFile(file):
    wb = openpyxl.load_workbook(file)
    ws = wb.get_active_sheet()
    result = False
    matchRowIndex = 0
    for cell in ws['A']:
        if cell.value == firstColumn.value:
            matchRowIndex = cell.row
            tmpIndex = 0
            bookCols = ws[matchRowIndex]
            if len(mergeColumns) != len(bookCols):
                return (False, matchRowIndex)
            for rowCell in bookCols:
                if(rowCell.value != mergeColumns[tmpIndex].value):
                    return (False, matchRowIndex)
                tmpIndex = tmpIndex+1
            result = True
    wb.close()
    return (result,matchRowIndex)

def replaceBreedNames(breedCols):
    for breedCell in breedCols:
        matchedStart = False
        for breedKey in breedMap.keys():
            if str(breedCell.value).startswith(breedKey):
                matchedStart = True
                breedCell.value = breedMap[breedKey]
        if not matchedStart:
            breedCell.value = breedMap['default']

def convertFullRowsToObject(fullSheet):
    fullRows = fullSheet[2:fullSheet.max_row]
    objectRecords = []
    for fullRow in fullRows:
        convertedObject = FullRecord()
        convertedObject.id = fullRow[0].value
        convertedObject.age = fullRow[1].value
        convertedObject.breed = fullRow[2].value
        convertedObject.status = fullRow[3].value
        convertedObject.breedingStatus = fullRow[4].value
        convertedObject.sinceLastAI = fullRow[5].value
        convertedObject.daysInMilk = fullRow[6].value
        convertedObject.toConception = fullRow[7].value
        convertedObject.avgMilk = fullRow[8].value
        convertedObject.routeName = fullRow[9].value
        convertedObject.hmbName = fullRow[10].value
        objectRecords.append(convertedObject)
    return objectRecords

def mergeFiles():
    print('Files for merging:')
    fileList = os.listdir(os.path.join('..', 'InputForMerge'))
    print(fileList)

    for file in fileList:
        (validationResult, rowIndex) = validateFile(os.path.join('..', 'InputForMerge', file))
        print("Validating file :", file, " - Result:", validationResult)
        if validationResult != True:
            print("This file will not be merged. Please check whether the file is having same number/name columns")
        else:
            print("Merging File :", file)
            wb = openpyxl.load_workbook(os.path.join('..', 'InputForMerge', file))
            ws = wb.get_active_sheet()
            rowsForMerge = ws[rowIndex + 1:ws.max_row]
            for newRow in rowsForMerge:
                rowVals = []
                for col in newRow:
                    rowVals.append(col.value)
                outputSheet.append(rowVals)

    breedColumns = outputSheet['I'][1:]
    replaceBreedNames(breedColumns)
    outputBook.save(os.path.join('..', 'Output', 'File.xlsx'))
    input("Finished...Press a key to exit")


def doHMBProcessing():
    print('Renaming Breed columns for you...')
    for hmbFile in os.listdir(os.path.join('..', 'InputForHMB')):
        hmbWorkbook = openpyxl.load_workbook(os.path.join('..', 'InputForHMB', hmbFile))
        fullWorksheet = hmbWorkbook.get_sheet_by_name('FULL')
        hmbSheet = hmbWorkbook.get_sheet_by_name('HMB wise')
        breedColumns = outputSheet['I'][1:]
        replaceBreedNames(breedColumns)
        fullRecords = convertFullRowsToObject(fullWorksheet)
        hmbIndexedMap = {}
        hmbConsolidatedMap = {}
        for fullRecord in fullRecords:
            if(fullRecord.hmbName in hmbIndexedMap):
                hmbIndexedMap[fullRecord.hmbName].append(fullRecord)
            else:
                indexedList  = []
                indexedList.append(fullRecord)
                hmbIndexedMap[fullRecord.hmbName] = indexedList

        for location in hmbIndexedMap.keys():
            locationRecords = hmbIndexedMap[location]
            for locRecord in locationRecords:
                if locRecord.status == 'Milking':
                    if locRecord.hmbName not in hmbConsolidatedMap:
                        newHmbType = HMBType()
                        newHmbType.bred = BreedType()
                        newHmbType.unbred = BreedType()
                        newHmbType.pregnant = BreedType()
                        hmbConsolidatedMap[location] = newHmbType
                    if locRecord.breedingStatus == 'Open Unbred':
                        hmbConsolidatedMap[location].unbred.total = (hmbConsolidatedMap[location].unbred.total)+1
                        if locRecord.status == 'HF':
                            hmbConsolidatedMap[location].unbred.HF = (hmbConsolidatedMap[
                                    location].unbred.HF) + 1
                        elif locRecord.status == 'Jersey':
                                hmbConsolidatedMap[location].unbred.Jy = (hmbConsolidatedMap[
                                    location].unbred.Jy) + 1
                        else:
                                hmbConsolidatedMap[location].unbred.ND = (hmbConsolidatedMap[
                                    location].unbred.ND) + 1
                    elif locRecord.breedingStatus == 'Open Bred':
                        hmbConsolidatedMap[location].bred.total = (hmbConsolidatedMap[location].bred.total) + 1
                        if locRecord.status == 'HF':
                                hmbConsolidatedMap[location].bred.HF = (hmbConsolidatedMap[
                                    location].bred.HF) + 1
                        elif locRecord.status == 'Jersey':
                                hmbConsolidatedMap[location].bred.Jy = (hmbConsolidatedMap[
                                    location].bred.Jy) + 1
                        else:
                                hmbConsolidatedMap[location].bred.ND = (hmbConsolidatedMap[
                                    location].bred.ND) + 1
                    elif locRecord.breedingStatus == 'Pregnant':
                        hmbConsolidatedMap[location].pregnant.total = (hmbConsolidatedMap[location].pregnant.total) + 1
                        if locRecord.status == 'HF':
                                hmbConsolidatedMap[location].pregnant.HF = (hmbConsolidatedMap[
                                    location].pregnant.HF) + 1
                        elif locRecord.status == 'Jersey':
                                hmbConsolidatedMap[location].pregnant.Jy = (hmbConsolidatedMap[
                                    location].pregnant.Jy) + 1
                        else:
                                hmbConsolidatedMap[location].pregnant.ND = (hmbConsolidatedMap[
                                    location].pregnant.ND) + 1
        hosurData = hmbConsolidatedMap['HOSUR']
        print(hosurData.unbred.total)
        print(hosurData.bred.total)
        print(hosurData.pregnant.total)

print('1: Merging')
print('2: HMB Processing')
optionval = input('Select your option: ')
if optionval == '1':
    mergeFiles()
elif optionval == '2':
    doHMBProcessing()
else:
    print('Wrong choice')